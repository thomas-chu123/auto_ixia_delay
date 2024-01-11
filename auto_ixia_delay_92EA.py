import Utils_9_2EA.IxLoadUtils as IxLoadUtils
import Utils_9_2EA.IxRestUtils as IxRestUtils
import Utils_9_2EA.IxLoadTestSettings as TestSettings
import inspect
import paramiko
import logging
import sys
import time
import openpyxl
import copy
from datetime import date
# import serial
import tkinter as tk
# import tkinter.filedialog as fd
import tkinter.messagebox as msgbox
from tkinter import ttk
# from ctypes import *
import ctypes
import ctypes.wintypes
from threading import Thread
# import subprocess
import os
import win32con
import statistics as stat
import csv

# ###TEST CONFIG
output_logging = 20
resetCard = True
runOnDifferentSetup = True
kGatewayServer = "192.168.42.37"
kGatewayPort = 8443  # TODO - to be changed by user depending on whether HTTP redirect is used or not
# TODO - to be changed by user in order to use HTTP requests instead of HTTPS requests (the HTTP requests will be redirected as HTTPS requests)
kHttpRedirect = False
kIxLoadVersion = "8.50.0.465"  # TODO - to be changed by user
# kRxfPath = "C:\\Tool\\auto_ixia_delay\\"
# kGatewaySharedFolder = 'C:\\ProgramData\\Ixia\\IxLoadGateway'  # TODO - to be changed by user depending on the gateway OS
# kRxfRelativeUploadPath = 'uploads\\%s' % os.path.split(kRxfPath)[1]  # TODO - to be changed by user
# kRxfAbsoluteUploadPath = '\\'.join([kGatewaySharedFolder, kRxfRelativeUploadPath])
FILE_LOCAL = ".\\Profile\\"
FILE_REMOTE = "C:\\ProgramData\\Ixia\\IxLoadGateway\\uploads\\"
kRxfPath = r"C:\\Path\\to\\config.rxf"
gatewayDiagnosticsFile = "gatewaydiags.zip"
diagsFile = "diags.zip"
gatewayDiagnosticsPath = '/'.join([os.path.dirname(kRxfPath), gatewayDiagnosticsFile])
diagsPath = '/'.join([os.path.dirname(kRxfPath), diagsFile])


kChassisList = ['192.168.42.5']  # TODO - to be changed by user
kApiKey = ''  # TODO - to be changed by user
kOS = "Windows"  # TODO - to be changed by user depending on the remote OS Windows/Linux
location = inspect.getfile(inspect.currentframe())
WindowsPath = ".\\Log\\"  # TODO - to be changed by user if you run the script on Windows machine
# This is the path where rxf file will be saved on the remote machine
kPortListPerCommunityCommunity = {
    #  format: { community name : [ port list ] }
    "Traffic1@Network1": [(1, 5, 3)],
    "Traffic2@Network2": [(1, 5, 4)]
}
kStatsToDisplayDict = {
    # format: { statSource : [stat name list] }
    "HTTPClient": ["HTTP Bytes Sent", "HTTP Bytes Received"],
    "HTTPServer": [""]
}
# http://192.168.42.34:8080/api/v0/sessions/35/ixload/test/activeTest/communityList/0/activityList
kActivityOptionsToChange = {
    # format: { activityName : { option : value } }
    "HTTPClient1": {
        "enableConstraint": True,
        "constraintType": "SimulatedUserConstraint",
        "constraintValue": 20,
        "userObjectiveType": "throughputMbps",
        "userObjectiveValue": 1000
    },
    "HTTPClient2": {
        "enableConstraint": True,
        "constraintType": "SimulatedUserConstraint",
        "constraintValue": 20,
        "userObjectiveType": "throughputMbps",
        "userObjectiveValue": 1000
    }
}
# http://192.168.42.34:8080/api/v0/sessions/38/ixload/test/activeTest/communityList/0/network/stack/childrenList/2/childrenList/3/rangeList
kIpOptionsToChange = {
    # format : { IP Range name : { optionName : optionValue } }
    "IP-R1": {
        "count": 1,
        "ipAddress": "10.1.1.187",
        "gatewayAddress": "10.1.1.188"
    },
    "IP-R2": {
        "count": 1,
        "ipAddress": "10.1.1.188",
        "gatewayAddress": "10.1.1.187"
    }
}

target_delay_pattern = ['0.0ms', '7.0ms', '15.0ms', '20.0ms']
real_delay_pattern = ['0.0ms', '2.0ms', '6.0ms', '8.0ms']
direction = ['download', 'upload', 'download_with_load', 'upload_with_load']
DELAY_IP_ADDR = "192.168.42.150"
DELAY_USERNAME = "root"
DELAY_PASSWORD = "123456"
vCPE_Console = "COM1 - Tera Term VT"
test_profile = ['100_40', '250_125', '1000_500', '1000_1000']
test_profile_HEX = {"100_40": "F2101213010331", "250_125": "F2101213010330", "1000_500": "F2101213010328",
                    "1000_1000": "F2101213010327"}
test_profile_ASCII = {"100_40": "1111111131", "250_125": "1111111130", "1000_500": "1111111128",
                      "1000_1000": "1111111129"}

DEVICE_TYPE = ['BRIDGE', 'ROUTER']

SIZE_ROW = 2
SIZE_COL = 4
START_ROW = 3
START_COL = 4
data_loc = {
    'C1': 0,
    'C2': 1 * SIZE_ROW,
    'C3': 2 * SIZE_ROW,
    'C4': 3 * SIZE_ROW,
    'C5': 4 * SIZE_ROW,
    'AVG': 5 * SIZE_ROW,
    '100_40': START_ROW,
    '250_125': START_ROW + SIZE_ROW * 6,
    '1000_500': START_ROW + SIZE_ROW * 12,
    '1000_1000': START_ROW + SIZE_ROW * 18,
    'download': 1,
    'upload': 2,
    'download_with_load': 3,
    'upload_with_load': 4,
    '0.0ms': START_COL,
    '7.0ms': START_COL + SIZE_COL * 1,
    '15.0ms': START_COL + SIZE_COL * 2,
    '20.0ms': START_COL + SIZE_COL * 3

}


# ###############
class UI(tk.Tk):
    def __init__(self):
        super().__init__()
        self.ixload_thread = 0
        self.proc_state = False
        self.report_name = ""
        x_loc = 0.001
        y_loc = 0.001
        x_gap = 0.13
        y_gap = 0.045

        self.start_logging()
        logging.info("Start Application")
        self.geometry("1000x700+25+25")
        self.title("IxLoad Delay Performance Automation v0.9 (2021/12/1)")

        y_loc += y_gap
        self.Port1_Label = tk.Label(self, text='IXIA Client Port:', anchor="w")
        self.Port1_Label.place(relx=x_loc, rely=y_loc, height=32, width=150)
        self.Port1 = tk.Entry(self, width=100)
        self.Port1.place(relx=x_loc + x_gap, rely=y_loc, relheight=0.03, relwidth=0.15)
        self.Port1.insert(tk.END, "1,5,3")

        y_loc += y_gap
        self.Port2_Label = tk.Label(self, text='IXIA Server Port:', anchor="w")
        self.Port2_Label.place(relx=x_loc, rely=y_loc, height=32, width=150)
        self.Port2 = tk.Entry(self, width=100)
        self.Port2.place(relx=x_loc + x_gap, rely=y_loc, relheight=0.03, relwidth=0.15)
        self.Port2.insert(tk.END, "1,5,4")

        y_loc += y_gap
        self.LanIP_Label = tk.Label(self, text='IXIA LAN IP:', anchor="w")
        self.LanIP_Label.place(relx=x_loc, rely=y_loc, height=20, width=150)
        self.LanIP = tk.Entry(self, width=100)
        self.LanIP.place(relx=x_loc + x_gap, rely=y_loc, relheight=0.03, relwidth=0.15)
        self.LanIP.insert(tk.END, "10.1.1.187")

        y_loc += y_gap
        self.GatewayIP_Label = tk.Label(self, text='IXIA Gateway IP:', anchor="w")
        self.GatewayIP_Label.place(relx=x_loc, rely=y_loc, height=20, width=150)
        self.GatewayIP = tk.Entry(self, width=100)
        self.GatewayIP.place(relx=x_loc + x_gap, rely=y_loc, relheight=0.03, relwidth=0.15)
        self.GatewayIP.insert(tk.END, "10.1.1.188")

        y_loc += y_gap
        self.DeviceType_Label = tk.Label(self, text='Device Type:', anchor="w")
        self.DeviceType_Label.place(relx=x_loc, rely=y_loc, height=20, width=150)
        self.DeviceType = ttk.Combobox(self, width=100, values=DEVICE_TYPE)
        self.DeviceType.current(0)
        self.DeviceType.place(relx=x_loc + x_gap, rely=y_loc, relheight=0.03, relwidth=0.15)
        self.DeviceType.bind("<<ComboboxSelected>>", self.device_change)

        y_loc += y_gap
        self.LanVLAN_Label = tk.Label(self, text='LAN VLAN:', anchor="w")
        self.LanVLAN_Label.place(relx=x_loc, rely=y_loc, height=20, width=150)
        self.LanVLAN = tk.Entry(self, width=100)
        self.LanVLAN.place(relx=x_loc + x_gap, rely=y_loc, relheight=0.03, relwidth=0.15)
        self.LanVLAN.insert(tk.END, "1")

        y_loc += y_gap
        self.ProfileList = []
        self.Profile_Label = tk.Label(self, text='WAN Profile:', anchor="w")
        self.Profile_Label.place(relx=x_loc, rely=y_loc, height=20, width=150)
        for id_name in range(0, 4):
            y_loc += y_gap * 0.8
            self.ProfileCheck = ""
            self.ProfileCheck = ttk.Checkbutton(self, text=test_profile[id_name])
            self.ProfileCheck.place(relx=x_loc + x_gap, rely=y_loc, height=32, width=150)
            self.ProfileCheck.state(['!alternate'])
            self.ProfileCheck.state(['selected'])
            self.ProfileList.append(self.ProfileCheck)

        y_loc += y_gap
        self.DelayList = []
        self.Delay_Label = tk.Label(self, text='Impairment/Delay:', anchor="w")
        self.Delay_Label.place(relx=x_loc, rely=y_loc, height=20, width=150)
        for id_name in range(0, 4):
            y_loc += y_gap * 0.8
            self.DelayCheck = ""
            self.DelayCheck = ttk.Checkbutton(self, text=target_delay_pattern[id_name])
            self.DelayCheck.place(relx=x_loc + x_gap, rely=y_loc, height=32, width=150)
            self.DelayCheck.state(['!alternate'])
            self.DelayCheck.state(['selected'])
            self.DelayList.append(self.DelayCheck)

        y_loc += y_gap
        self.DirectionList = []
        self.Direction_Label = tk.Label(self, text='Traffic Direction:', anchor="w")
        self.Direction_Label.place(relx=x_loc, rely=y_loc, height=20, width=150)
        for id_name in range(0, 4):
            y_loc += y_gap * 0.8
            self.DirectionCheck = ""
            self.DirectionCheck = ttk.Checkbutton(self, text=direction[id_name])
            self.DirectionCheck.place(relx=x_loc + x_gap, rely=y_loc, height=32, width=200)
            self.DirectionCheck.state(['!alternate'])
            self.DirectionCheck.state(['selected'])
            self.DirectionList.append(self.DirectionCheck)

        y_loc += y_gap
        self.TestCount_Label = tk.Label(self, text='Test Count:', anchor="w")
        self.TestCount_Label.place(relx=x_loc, rely=y_loc, height=20, width=150)
        self.TestCount = ttk.Combobox(self, width=100, values=[1, 2, 3, 4, 5])
        self.TestCount.current(0)
        self.TestCount.place(relx=x_loc + x_gap, rely=y_loc, relheight=0.03, relwidth=0.15)
        # self.DeviceType.bind("<<ComboboxSelected>>", self.test_profile_change)

        x_loc = 0.01
        y_loc += y_gap + 0.015
        self.StartButton = tk.Button(self, pady="0", text='Start Test', command=self.start_test)
        self.StartButton.place(relx=x_loc, rely=y_loc, height=31, width=150)
        self.StartButton['state'] = tk.NORMAL

        x_loc = 0.15
        y_loc = y_loc
        self.StopButton = tk.Button(self, pady="0", text='Exit Test', command=self.stop_test)
        self.StopButton.place(relx=x_loc, rely=y_loc, height=31, width=150)
        self.StopButton['state'] = tk.DISABLED

        x_loc = 0.32
        y_loc = 0.01
        self.Text_Label = tk.Label(self, text='Output:', anchor="w")
        self.Text_Label.place(relx=x_loc, rely=y_loc, height=20, width=150)
        y_loc += y_gap
        self.ResponseText = tk.Text(self, font=("Arial", 10))
        self.ResponseText.place(relx=x_loc, rely=y_loc, height=600, width=660)
        self.ScrollBar = tk.Scrollbar(self.ResponseText)
        self.ScrollBar.pack(side=tk.RIGHT, fill=tk.Y)
        self.ScrollBar.config(command=self.ResponseText.yview)

    def device_change(self, event):
        self.LanIP.delete(0, tk.END)
        self.GatewayIP.delete(0, tk.END)
        self.LanVLAN.delete(0, tk.END)
        if self.DeviceType.get() == "BRIDGE":
            self.LanIP.insert(tk.END, "10.1.1.187")
            self.GatewayIP.insert(tk.END, "10.1.1.188")
            self.LanVLAN.insert(tk.END, "7")
        else:
            self.LanIP.insert(tk.END, "192.168.1.33")
            self.GatewayIP.insert(tk.END, "192.168.1.1")
            self.LanVLAN.insert(tk.END, "1")

    def start_test(self):
        try:
            logging.info("Start IxLoad Testing")
            self.report_name = ""
            self.StartButton['state'] = tk.DISABLED
            self.StopButton['state'] = tk.NORMAL

            self.ixia_port1 = self.Port1.get()
            self.ixia_port2 = self.Port2.get()
            self.dut_type = self.DeviceType.get()
            self.lan_ip = self.LanIP.get()
            self.gateway_ip = self.GatewayIP.get()
            self.lan_vlan = int(self.LanVLAN.get())
            resetCard=True

            for profile in self.ProfileList:
                if 'selected' in profile.state():
                    self.dut_profile = profile.cget("text")
                    print(self.dut_profile)
                    logging.info("Start Testing Profile %s ", self.dut_profile)
                    self.update_output("Test Profile: " + self.dut_profile)
                    # Configure CPE and Wait CPE to bring up PON
                    logging.info("Start Configure CPE Profile %s ", self.dut_profile)
                    self.update_output("Configure CPE Console: " + self.dut_profile)
                    self.update_output("Waiting 100 seconds for Rebooting CPE.....")
                    self.change_cpe_profile(self.dut_type)
                    # Configure delay server
                    for delay in self.DelayList:
                        if 'selected' in delay.state():
                            self.dut_delay = delay.cget("text")
                            print(self.dut_delay)
                            logging.info("Start Configure Delay Server %s ", self.dut_delay)
                            self.update_output("Configure Delay Server: " + self.dut_delay)
                            for id_name in range(len(target_delay_pattern)):
                                if target_delay_pattern[id_name] == self.dut_delay:
                                    self.delay_time = real_delay_pattern[id_name]
                            self.delay_control(self.delay_time)
                            # Configure direction
                            for direction in self.DirectionList:
                                if 'selected' in direction.state():
                                    self.dut_direction = direction.cget("text")
                                    logging.info("Start Test Direction %s ", self.dut_direction)
                                    self.update_output("Test Direction: " + self.dut_direction)
                                    #print(self.dut_direction)
                                    for count in range(int(self.TestCount.get())):
                                        # For real running
                                        logging.info("Test Count: %s ", str(count + 1))
                                        logging.info("Start IXLOAD process: %s ", str(count + 1))
                                        self.update_output("Test Count: " + str(count + 1))
                                        self.update_output("Start IXLOAD: " + str(count + 1))
                                        self.ixload_thread = Ixload_Test(self, self.ixia_port1, self.ixia_port2,
                                                                         self.dut_type,
                                                                         self.dut_direction, self.lan_ip, self.gateway_ip,
                                                                         self.dut_profile, self.lan_vlan)
                                        result_folder = self.ixload_thread.result_folder
                                        self.update_output("Parser Test Result: " + result_folder + "/HTTP_Client.csv")
                                        data_result = self.parser_data(result_folder + "/HTTP_Client.csv")
                                        logging.info("Start Parser data: %s ", result_folder + "/HTTP_Client.csv")
                                        self.save_csv(result_folder + "/HTTP_Client.csv", self.dut_direction,
                                                      self.dut_profile, self.dut_delay, count)

                                        # For testing only
                                        # data_result = self.parser_data("HTTP_Client.csv")
                                        # self.save_csv("HTTP_Client.csv", self.dut_direction,
                                        #              self.dut_profile, self.dut_delay, count)

                                        self.update_output("Write Data to Test Report")
                                        self.update_output("Download Throughput (Mbps): " + str(data_result[2]))
                                        self.update_output("Download STDEV: " + str(data_result[3]))
                                        self.update_output("Upload Throughput (Mbps): " + str(data_result[0]))
                                        self.update_output("Upload STDEV: " + str(data_result[1]))

                                        logging.info("Download: %s ", str(data_result[2]))
                                        logging.info("Upload: %s ", str(data_result[0]))

                                        self.generate_report("test_report_template.xlsx", count, self.dut_profile,
                                                             self.dut_delay, self.dut_direction, data_result)
                                        self.update_output("Wait 10s for Next Round")
                                        logging.info("Write Data to Report and wait 10s for next round")
                                        self.update_output("\r\n")

                                        for i in range(20):
                                            self.update()
                                            self.update_idletasks()
                                            app.update()
                                            self.after(500, "")

            msgbox.showinfo("INFO", "IxLoad Test Finished")
            self.StartButton['state'] = tk.NORMAL
            self.StopButton['state'] = tk.DISABLED
        except Exception as ex:
            print("Test Control is failed and Exception Occur:",repr(ex))

    def stop_test(self):
        try:
            self.ixload_thread.ixia_stop()
            self.StartButton['state'] = tk.NORMAL
            self.StopButton['state'] = tk.DISABLED
            sys.exit(0)
        except:
            sys.exit(0)

    def delay_control(self, delay_ms):
        try:
            for i in range(10):
                self.ssh_connect(DELAY_IP_ADDR, DELAY_USERNAME, DELAY_PASSWORD, "/opt/ixload_auto/delay.sh " + delay_ms)
                time.sleep(3)
                delay_result = self.ssh_connect(DELAY_IP_ADDR, DELAY_USERNAME, DELAY_PASSWORD, "tc qdisc |grep netem")

                if delay_ms == "0.0ms":
                    if "delay" not in delay_result:
                        logging.info("Configure delay success: %s", delay_ms)
                        break
                    else:
                        logging.info("Configure delay fail: %s", delay_ms)
                        continue
                else:
                    if delay_ms in delay_result:
                        logging.info("Configure delay success: %s", delay_ms)
                        break
                    else:
                        logging.info("Configure delay fail: %s", delay_ms)
                        continue
        except Exception as ex:
            print("Delay Control is failed and Exception Occur:",repr(ex))

    def ssh_connect(self, ip_addr, user, password, command):
        ssh_client = paramiko.SSHClient()
        ssh_client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
        stdin = ''
        stdout = ''
        stderr = ''
        stdout_output = ""
        stderr_output = ""
        try:
            # logging.info("Connect to SSH server: " + ip_addr)
            ssh_client.connect(hostname=ip_addr, port=22, username=user, password=password, timeout=10,
                               auth_timeout=10, banner_timeout=10)
            # stdin, stdout, stderr = ssh_client.exec_command(command=command, timeout=10)
            channel = ssh_client.get_transport().open_session()
            channel.exec_command(command=command)
            channel.shutdown_write()
            exit_code = channel.recv_exit_status()

            if "cat" in command or "tc qdisc" in command:
                stdout = channel.makefile().read().decode()
                # stderr = channel.makefile_stderr().read().decode()
                stdout_output = stdout

            # logging.info("Err: " + stderr_output + "\r\n")
            # logging.info("Out: " + stdout_output + "\r\n")
            channel.close()
            ssh_client.close()
            return stdout_output
        except Exception as e:
            logging.info("Connect Server SSH fail: " + repr(e))
            print("Exception: ", repr(e))
            ssh_client.close()

    def generate_report(self, template_file, test_id, profile, delay, direction, result):
        try:
            if self.report_name == "":
                workbook = openpyxl.load_workbook(FILE_LOCAL + template_file)
                worksheet = workbook['result']
                today = date.today()
                fw_name = "test_report_" + today.strftime("%Y_%m_%d") + ".xlsx"
                self.report_name = fw_name
                workbook.save(fw_name)
                workbook.close()

            workbook = openpyxl.load_workbook(self.report_name)
            worksheet = workbook['result']

            x_cell = data_loc[delay] + data_loc[direction] - 1
            y_cell = data_loc[profile] + data_loc['C' + str(test_id + 1)]

            if 'download' in direction:
                worksheet.cell(y_cell, x_cell, ).value = result[2]
                worksheet.cell(y_cell + 1, x_cell, ).value = result[3]
            else:
                worksheet.cell(y_cell, x_cell, ).value = result[0]
                worksheet.cell(y_cell + 1, x_cell, ).value = result[1]

            workbook.save(self.report_name)
            workbook.close()
        except Exception as ex:
            print("Generate Report is failed and Exception Occur:",repr(ex))

    def csv_read(self, filename):
        try:
            table = []
            logging.info('open filename = ' + filename)
            file = open(filename, mode="r", errors='ignore')
            table = file.readlines()
            file.close()
            return table
        except Exception as ex:
            print("Read CSV is failed and Exception Occur:", repr(ex))

    def save_csv(self, filename, direction, profile, delay, count):
        try:
            data = self.csv_read(filename)
            file = open(WindowsPath + direction + '_' + profile + '_' + delay + '_' + str(count + 1) + '.csv', mode="w",
                        errors='ignore')
            file.writelines(data)
            file.close()
        except Exception as ex:
            print("Write CSV is failed and Exception Occur:",repr(ex))

    def parser_data(self, report_file):
        try:
            data = self.csv_read(report_file)
            # calculate tx_rate
            tx_rate = 0
            rx_rate = 0
            tx_rate_sum = 0
            rx_rate_sum = 0
            tx_rate_avg = 0
            rx_rate_avg = 0
            tx_rate_std = 0
            rx_rate_std = 0
            tx_rate_list = []
            rx_rate_list = []
            x_cell = 112
            y_cell = 23
            print("Row, Tx_Rate (kbps), Rx_Rate (Kbps)")
            for row in range(y_cell - 1, y_cell + 23):
                tx_rate = float(data[row].split(',')[x_cell - 1])
                tx_rate_sum = tx_rate_sum + tx_rate
                tx_rate_list.append(tx_rate)
                rx_rate = float(data[row].split(',')[x_cell])
                rx_rate_sum = rx_rate_sum + rx_rate
                rx_rate_list.append(rx_rate)
                print(row + 1, ",", tx_rate, ",", rx_rate)
            print("AVG:")
            tx_rate_avg = round(tx_rate_sum / 24 / 1000, 2)
            rx_rate_avg = round(rx_rate_sum / 24 / 1000, 2)
            print(tx_rate_sum / 24, ",", rx_rate_sum / 24)
            print("STDEV:")
            tx_rate_std = round(stat.stdev(tx_rate_list) / 1000, 2)
            rx_rate_std = round(stat.stdev(rx_rate_list) / 1000, 2)
            print(tx_rate_std, ",", rx_rate_std)
            return [tx_rate_avg, tx_rate_std, rx_rate_avg, rx_rate_std]
        except Exception as ex:
            print("Parser Data is failed and Exception Occur:",repr(ex))

    def start_logging(self):
        # Enable the logging to console and file
        for handler in logging.root.handlers[:]:
            logging.root.removeHandler(handler)
        logging.basicConfig(level=output_logging,
                            format='%(asctime)s: [%(levelname)s] %(message)s',
                            datefmt='%a, %d %b %Y %H:%M:%S',
                            filename='ixload_test.log',
                            filemode='w')

        console = logging.StreamHandler()
        console.setLevel(output_logging)
        formatter = logging.Formatter('%(levelname)-4s %(message)s')
        console.setFormatter(formatter)
        logging.getLogger('').addHandler(console)

    def update_output(self, msg):
        self.update()
        self.update_idletasks()
        self.ResponseText.insert(tk.END, msg + "\r\n")
        self.ResponseText.focus_set()
        self.ResponseText.focus_lastfor()
        self.ResponseText.see(tk.END)
        self.update()
        self.update_idletasks()

    def change_cpe_profile(self, dut_type):
        try:
            if dut_type=="ROUTER":
                command_list = [b"\r\ndiag gpon set tx-laser force-off\r\n",
                                b"\r\ntefapp set ont ploam_password " + test_profile_HEX[self.dut_profile].encode() + b" \r\n",
                                b"\r\ndiag gpon set tx-laser normal\r\n",
                                b"\r\nreboot\r\n",
                                b"\r\n",
                                b"\r\n",
                                b"1234\r\n",
                                b"12345678\r\n",
                                b"\r\n",
                                b"\r\nsh\r\n"]
            else:
                command_list = [b"\r\ndiag gpon set tx-laser force-off\r\n",
                                b"\r\ntefapp set ont ploam_password " + test_profile_HEX[self.dut_profile].encode() + b" \r\n",
                                b"\r\ndiag gpon set tx-laser normal\r\n",
                                b"\r\nreboot\r\n",
                                b"\r\n",
                                b"\r\n",
                                b"admin\r\n",
                                b"12345678\r\n",
                                b"\r\n",
                                b"\r\nsh\r\n"]

            for cmd in command_list:
                self.set_thread = set_cpe2(vCPE_Console, cmd)
                if b"\r\nreboot\r\n" in cmd:
                    time_delay = 200
                else:
                    time_delay = 6
                for i in range(time_delay):
                    self.update()
                    self.update_idletasks()
                    app.update()
                    self.after(500, "")
        except Exception as ex:
            print("Change CPE profile is failed and Exception Occur:",repr(ex))

class Ixload_Test(Thread):
    def __init__(self, ui_object, port1, port2, device_type, direction, lan_ip, gateway_ip, profile, vlan):
        super().__init__()

        self.device_type = device_type
        self.direction = direction
        self.port1 = port1
        self.port2 = port2
        self.lan_ip = lan_ip
        self.gateway_ip = gateway_ip
        self.ui_object = ui_object
        self.profile = profile
        self.vlan = vlan
        self.ixia_start()

    def ixia_start(self):
        global resetCard
        # create a connection to the gateway
        self.testSettings = TestSettings.IxLoadTestSettings()
        self.connection = IxRestUtils.getConnection(self.testSettings.gatewayServer, self.testSettings.gatewayPort,
            httpRedirect=self.testSettings.httpRedirect, version=self.testSettings.apiVersion)
        self.connection.setApiKey(self.testSettings.apiKey)
        self.sessionUrl = None
        # create a session
        try:
            IxLoadUtils.log("Creating a new session...")
            self.ui_object.update_output("Creating a new session...")
            sessionUrl = IxLoadUtils.createNewSession(self.connection, self.testSettings.ixLoadVersion)
            IxLoadUtils.log("Session created.")
            self.ui_object.update_output("Session created.")
            if resetCard == True:
                IxLoadUtils.enableForcefullyTakeOwnershipAndResetPorts(self.connection, self.sessionUrl)
                IxLoadUtils.log("Wait 5min for IXIA to take ownership and reset port")
                self.ui_object.update_output("Wait 5min for IXIA to take ownership and reset port")

                for i in range(10):
                    self.ui_object.update()
                    self.ui_object.update_idletasks()
                    self.ui_object.after(500, "")

            filename = self.direction + ".rxf"
            # upload file to gateway server
            if not self.testSettings.isLocalHost():
                IxLoadUtils.log('Uploading file %s...' % filename)
                self.kResourcesUrl = IxLoadUtils.getResourcesUrl(self.connection)
                IxLoadUtils.uploadFile(self.connection, self.kResourcesUrl, FILE_LOCAL + filename, "uploads\\" + filename)
                IxLoadUtils.log('Upload file finished.')
                self.ui_object.update_output("Upload file finished.")

            # load a repository
            IxLoadUtils.log("Loading repository %s" % (FILE_REMOTE + filename))
            IxLoadUtils.loadRepository(self.connection, self.sessionUrl, FILE_REMOTE + filename)
            IxLoadUtils.log("Repository loaded.")
            self.ui_object.update_output("Repository loaded.")

            # Modify CSVs results path
            # load_test_url = "%s/ixload/test/" % self.session_url
            # payloadDict = {"outputDir" : "true", "runResultDirFull" : csv_path}
            # IxLoadUtils.log("Perform CSVs results path modification.")
            # IxLoadUtils.performGenericPatch(self.connection, load_test_url, payloadDict)
            if runOnDifferentSetup:
                IxLoadUtils.log("Clearing chassis list...")
                IxLoadUtils.clearChassisList(self.connection, self.sessionUrl)
                IxLoadUtils.log("Chassis list cleared.")
                self.ui_object.update_output("Chassis list cleared.")

                IxLoadUtils.log("Adding chassis %s..." % (kChassisList))
                IxLoadUtils.addChassisList(self.connection, sessionUrl, self.testSettings.chassisList)
                IxLoadUtils.log("Chassis added.")
                self.ui_object.update_output("Chassis added")

                port_list = copy.deepcopy(kPortListPerCommunityCommunity)
                port_list['Traffic1@Network1'] = [eval(self.port1)]
                port_list['Traffic2@Network2'] = [eval(self.port2)]

                IxLoadUtils.log("Assigning new ports...")
                IxLoadUtils.assignPorts(self.connection, self.sessionUrl, port_list)
                IxLoadUtils.assignPorts(self.connection, sessionUrl, self.testSettings.portListPerCommunity)
                IxLoadUtils.log("Ports assigned.")
                self.ui_object.update_output("Ports assigned")

                ip_option = copy.deepcopy(kIpOptionsToChange)
                if self.device_type == "BRIDGE":
                    ip_option["IP-R1"]["ipAddress"] = self.lan_ip
                    ip_option["IP-R1"]["gatewayAddress"] = self.gateway_ip
                    ip_option["IP-R2"]["ipAddress"] = "10.1.1.188"
                    ip_option["IP-R2"]["gatewayAddress"] = "10.1.1.187"
                else:
                    ip_option["IP-R1"]["ipAddress"] = self.lan_ip
                    ip_option["IP-R1"]["gatewayAddress"] = self.gateway_ip
                    ip_option["IP-R2"]["ipAddress"] = "10.1.1.188"
                    ip_option["IP-R2"]["gatewayAddress"] = "10.1.1.187"

                IxLoadUtils.log("Updating IP Ranges...")
                IxLoadUtils.changeIpRangesParams(self.connection, self.sessionUrl, ip_option)
                IxLoadUtils.log("IP Ranges updated.")
                self.ui_object.update_output("IP Ranges updated.")

                if self.vlan!=1:
                    IxLoadUtils.log("Updating VLAN Ranges...")
                    IxLoadUtils.changeVlanRangesParams(self.connection, self.sessionUrl, self.vlan)
                    IxLoadUtils.log("VLAN Ranges updated.")
                    self.ui_object.update_output("VLAN Ranges updated.")

                user_option = copy.deepcopy(kActivityOptionsToChange)
                if self.direction == "download":
                    if self.device_type == "BRIDGE":
                        user_option['HTTPClient1']['constraintValue'] = 12
                    else:
                        user_option['HTTPClient1']['constraintValue'] = 20
                elif self.direction == "upload":
                    if self.device_type == "BRIDGE":
                        user_option['HTTPClient2']['constraintValue'] = 12
                    else:
                        user_option['HTTPClient2']['constraintValue'] = 20
                elif self.direction == "download_with_load":
                    if self.device_type == "BRIDGE":
                        user_option['HTTPClient1']['constraintValue'] = 12
                        user_option['HTTPClient2']['constraintValue'] = 4
                    else:
                        user_option['HTTPClient1']['constraintValue'] = 20
                        user_option['HTTPClient2']['constraintValue'] = 4
                elif self.direction == "upload_with_load":
                    if self.device_type == "BRIDGE":
                        user_option['HTTPClient1']['constraintValue'] = 4
                        user_option['HTTPClient2']['constraintValue'] = 12
                    else:
                        user_option['HTTPClient1']['constraintValue'] = 4
                        user_option['HTTPClient2']['constraintValue'] = 20

                down_rate, up_rate = self.profile.split("_")

                user_option['HTTPClient1']['userObjectiveValue'] = int(down_rate)
                user_option['HTTPClient2']['userObjectiveValue'] = int(up_rate)

                IxLoadUtils.log("Updating activity options...")
                IxLoadUtils.changeActivityOptions(self.connection, self.sessionUrl, user_option)
                IxLoadUtils.log("Updated activity options.")
                self.ui_object.update_output("Updated activity options.")

                IxLoadUtils.log("Saving repository %s..." % (IxLoadUtils.getRxfName(kOS, location, WindowsPath)))
                self.filename = IxLoadUtils.getRxfName(kOS, location, WindowsPath)
                IxLoadUtils.saveRxf(self.connection, self.sessionUrl, self.filename)
                IxLoadUtils.log("Repository saved.")
                self.ui_object.update_output("Repository saved.")

            IxLoadUtils.log("Starting the test...")
            IxLoadUtils.runTest(self.connection, self.sessionUrl)
            IxLoadUtils.log("Test started.")
            self.ui_object.update_output("Test started.")

            IxLoadUtils.log("Polling values for stats %s..." % (kStatsToDisplayDict))
            IxLoadUtils.pollStats(self.ui_object, self.connection, self.sessionUrl, kStatsToDisplayDict)
            IxLoadUtils.log("Test finished.")
            self.ui_object.update_output("Test finished.")

            IxLoadUtils.log("Get Test Report Folder")
            self.result_folder = IxLoadUtils.getResultFolderUrl(self.connection, self.sessionUrl)
            IxLoadUtils.log(self.result_folder)
            self.ui_object.update_output("Get Test Report Folder")

            IxLoadUtils.log("Checking test status...")
            testRunError = IxLoadUtils.getTestRunError(self.connection, self.sessionUrl)
            if testRunError:
                IxLoadUtils.log("The test exited with the following error: %s" % testRunError)

                IxLoadUtils.log("Waiting for gateway diagnostics collection...")
                IxLoadUtils.collectGatewayDiagnostics(self.connection, gatewayDiagnosticsPath)
                IxLoadUtils.log("Gateway diagnostics are saved in %s" % gatewayDiagnosticsPath)

                IxLoadUtils.log("Waiting for diagnostics collection...")
                IxLoadUtils.collectDiagnostics(self.connection, self.sessionUrl, diagsPath)
                IxLoadUtils.log("Diagnostics are saved in %s" % diagsPath)

            else:
                IxLoadUtils.log("The test completed successfully.")

            IxLoadUtils.log("Waiting for test to clean up and reach 'Unconfigured' state...")
            IxLoadUtils.waitForTestToReachUnconfiguredState(self.connection, self.sessionUrl)
            IxLoadUtils.log("Test is back in 'Unconfigured' state.")

        except Exception as e:
            print("Error: " + repr(e))
            self.ui_object.update_output("Runtime Error: " + repr(e))

        finally:
            resetCard = False
            if self.sessionUrl is not None:
                IxLoadUtils.log("Closing IxLoad session...")
                IxLoadUtils.deleteSession(self.connection, self.sessionUrl)
                IxLoadUtils.log("IxLoad session closed.")
                self.ui_object.update_output("IxLoad session closed")


    def ixia_stop(self):
        if self.sessionUrl is not None:
            IxLoadUtils.log("Closing IxLoad session...")
            IxLoadUtils.deleteSession(self.connection, self.sessionUrl)
            IxLoadUtils.log("IxLoad session closed.")
            self.ui_object.update_output("Session created.")
        sys.exit(0)

class COPYDATASTRUCT(ctypes.Structure):
    _fields_ = [
        ('dwData', ctypes.wintypes.LPARAM),
        ('cbData', ctypes.wintypes.DWORD),
        ('lpData', ctypes.c_char_p)
        # formally lpData is c_void_p, but we do it this way for convenience
    ]

class set_cpe2():
    def __init__(self, title, msg):
        super().__init__()
        self.task = 0
        self.title = title
        self.msg = msg
        # self.daemon = True
        # self.start()
        self.run()

    def run(self):
        # prepare data and send it
        FindWindow = ctypes.windll.user32.FindWindowW
        SendMessage = ctypes.windll.user32.SendMessageW
        try:
            hwnd = FindWindow("VTWin32", self.title)
            cds = COPYDATASTRUCT()
            cds.dwData = 1
            cds.cbData = ctypes.sizeof(ctypes.create_string_buffer(self.msg))
            cds.lpData = ctypes.c_char_p(self.msg)
            SendMessage(hwnd, win32con.WM_COPYDATA, 0, ctypes.byref(cds))
        except Exception as e:
            logging.info("Write command to console failed, " + repr(e))

if __name__ == '__main__':
    app = UI()
    app.mainloop()
    sys.exit(0)
